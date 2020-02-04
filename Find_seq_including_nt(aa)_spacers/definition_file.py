class Y():   
   
   read_file_pass=input("Input the file pass including the read sequence .txt or .fasta files>")
   result_file_pass=input("Input the file pass for saving result .xlsx files>")

   def mkfile(self,name,sequence):
      self.name=name
      self.sequence=sequence
      from input_data_storage_file import Data_Storage
      input_name=Data_Storage.input_name
      input_element=Data_Storage.input_element
      input_spacer_number=Data_Storage.input_spacer_number

      import openpyxl
      from openpyxl import Workbook

      result_directory=Y.result_file_pass
         
      file_name="result_"+self.name+".xlsx"
      cordence_result_xlsx=r"{}".format(result_directory+"\cordence_"+file_name)
      discordence_result_xlsx=r"{}".format(result_directory+"\discordence_"+file_name)

      wb1=Workbook("test")
      wb2=Workbook("test")
      wb1.save(cordence_result_xlsx) 
      wb2.save(discordence_result_xlsx) 

      d=0        
      wb1 = openpyxl.load_workbook(cordence_result_xlsx)
      wb2 = openpyxl.load_workbook(discordence_result_xlsx)
      for d in range(len(input_name)):
         wb1.create_sheet(index=d,title=input_name[d])
         wb2.create_sheet(index=d,title=input_name[d])
         sheet1 = wb1[input_name[d]]
         sheet2 = wb2[input_name[d]]

         sheet1.cell(row=1, column=1,value="front bases"+":"+str(input_element[d][0])+"; "+str(len(input_element[d][0])))
         sheet1.cell(row=1, column=2,value="spacer bases"+";"+str(input_spacer_number[d]))
         sheet1.cell(row=1, column=3,value="back bases"+":"+str(input_element[d][1])+"; "+str(len(input_element[d][1])))
         sheet1.cell(row=1, column=4,value="region number")
         sheet1.cell(row=1, column=5,value="count")
         sheet1.cell(row=1, column=6,value="front bases"+":"+str(input_element[d][0])+"; "+str(len(input_element[d][0])))
         sheet1.cell(row=1, column=7,value="spacer bases"+";"+str(input_spacer_number[d]))
         sheet1.cell(row=1, column=8,value="back bases"+";"+str(input_element[d][1])+"; "+str(len(input_element[d][1])))
         sheet1.cell(row=1, column=9,value="region number")
         sheet1.cell(row=1, column=10,value="count")

         sheet2.cell(row=1, column=1,value="front bases"+":"+str(input_element[d][0])+"; "+str(len(input_element[d][0])))
         sheet2.cell(row=1, column=2,value="spacer bases"+";"+str(input_spacer_number[d]))
         sheet2.cell(row=1, column=3,value="expected back bases"+";"+str(len(input_element[d][1])))
         sheet2.cell(row=1, column=4,value="region number")
         sheet2.cell(row=1, column=5,value="count")
         sheet2.cell(row=1, column=6,value="expected front bases"+";"+str(len(input_element[d][0])))
         sheet2.cell(row=1, column=7,value="spacer bases"+";"+str(input_spacer_number[d]))
         sheet2.cell(row=1, column=8,value="back bases"+":"+str(input_element[d][1])+"; "+str(len(input_element[d][1])))
         sheet2.cell(row=1, column=9,value="region number")
         sheet2.cell(row=1, column=10,value="count")
         
         x=sequence.count(input_element[d][0])
         y=sequence.count(input_element[d][1])

         if x==0 and y==0:
            sheet1.cell(row=2, column=1,value=str(input_element[d][0])+" is NOT found")
            sheet1.cell(row=2, column=8,value=str(input_element[d][1])+" is NOT found")    
            sheet2.cell(row=2, column=1,value=str(input_element[d][0])+" is NOT found")
            sheet2.cell(row=2, column=8,value=str(input_element[d][1])+" is NOT found")  
            wb1.save(cordence_result_xlsx) 
            wb2.save(discordence_result_xlsx) 
            d=d+1
         elif x==0 and y!=0:
            sheet1.cell(row=2, column=1,value=str(input_element[d][0])+" is NOT found")
            sheet2.cell(row=2, column=1,value=str(input_element[d][0])+" is NOT found")
            wb1.save(cordence_result_xlsx) 
            wb2.save(discordence_result_xlsx) 
            d=d+1
         elif x!=0 and y==0:
            sheet1.cell(row=2, column=8,value=str(input_element[d][1])+" is NOT found")
            sheet2.cell(row=2, column=8,value=str(input_element[d][1])+" is NOT found")
            wb1.save(cordence_result_xlsx) 
            wb2.save(discordence_result_xlsx) 
            d=d+1
         elif x!=0 and y!=0:
            wb1.save(cordence_result_xlsx) 
            wb2.save(discordence_result_xlsx) 
            d=d+1
         else:
            wb1.save(cordence_result_xlsx) 
            wb2.save(discordence_result_xlsx) 
            d=d+1
               

   def Input_find(self,sequence,N_count):
       self.sequence=sequence
       self.N_count=N_count
       from input_data_storage_file import Data_Storage
       input_name=Data_Storage.input_name
       input_element=Data_Storage.input_element
       input_spacer_number=Data_Storage.input_spacer_number

       result_directory=Y.result_file_pass
      
       file_name="result_"+self.name+".xlsx"
       cordence_result_xlsx=r"{}".format(result_directory+"\cordence_"+file_name)
       discordence_result_xlsx=r"{}".format(result_directory+"\discordence_"+file_name)

       import openpyxl
       from openpyxl import Workbook
       wb1=openpyxl.load_workbook(cordence_result_xlsx)
       wb2=openpyxl.load_workbook(discordence_result_xlsx)
       
       from tqdm import tqdm
       
       v=0
       for v in tqdm(range(len(input_name))):
          import re
          print("\n")
          print("in "+input_element[v][0]+" turn for "+input_name[v])
          print("element"+" is "+input_element[v][0]+"/spacer sequences("+str(input_spacer_number[v])+")/"+input_element[v][1])  
          print("\n")
          iterator_front=re.finditer(input_element[v][0],self.sequence)  
          i=2
          j=2
          sheet1 = wb1[input_name[v]]
          sheet2 = wb2[input_name[v]] 
          wb1.save(cordence_result_xlsx) 
          wb2.save(discordence_result_xlsx) 
          s=sequence.find(input_element[v][0]) 

          spacer_num_width_v=input_spacer_number[v]
          snw=spacer_num_width_v
          len_input_element_v=len(input_element[v][1])
          lee=len_input_element_v

          if s==0:
             print(input_element[v][0]+" is not found")
             wb1.save(cordence_result_xlsx) 
             wb2.save(discordence_result_xlsx) 
             v=v+1 
          for m in iterator_front:
             a=m.group()
             b=m.start()+self.N_count
             c=m.end()+self.N_count    
             print(a,b,c)
             if sequence[c+snw:c+snw+lee]==input_element[v][1]:  
                print("cordence sequence exists")     
                sheet1.cell(row=i, column=1,value=sequence[b:c])
                sheet1.cell(row=i, column=2,value=sequence[c:c+snw])
                sheet1.cell(row=i, column=3,value=sequence[c+snw:c+snw+lee])
                sheet1.cell(row=i, column=4,value=str(b)+":"+str(c+snw+lee))
                sheet1.cell(row=i, column=5,value=sequence.count(sequence[b:c+snw+lee]))
                i=i+1
                try: 
                   pass
                except IndexError as IE:
                   print(IE)
                   v=v+1         
                except PermissionError as PE:
                   print(PE)
                   continue
                else:
                   continue                                                     
             elif sequence[c+snw:c+snw+lee]!=input_element[v][1]:
                #sheet2.cell(row=j, column=1,value=sequence[b:c])
                #sheet2.cell(row=j, column=2,value=sequence[c:c+snw])
                #sheet2.cell(row=j, column=3,value=sequence[c+snw:c+snw+lee])
                #sheet2.cell(row=j, column=4,value=str(b)+":"+str(c+snw+lee))
                #sheet2.cell(row=j, column=5,value=sequence.count(sequence[b:c+snw+lee]))
                j=j+1
                try: 
                   pass
                except IndexErroras as IE:
                   print(IE)
                   v=v+1            
                except PermissionError as PE:
                   print(PE)
                   continue
                else:
                   continue
             else:
                try: 
                   pass
                except IndexError as IE:
                   print(IE)
                   continue          
                except PermissionError as PE:
                   print(PE)
                   continue
                else:
                   continue
          else:
              wb1.save(cordence_result_xlsx) 
              wb2.save(discordence_result_xlsx)    
              v=v+1
 

         
       w=0
       for w in tqdm(range(len(input_name))):
          import re
          print("\n")
          print("in "+input_element[w][1]+" turn for "+input_name[w])
          print("element"+" is "+input_element[w][0]+"/spacer sequences("+str(input_spacer_number[w])+")/"+input_element[w][1])    
          print("\n")
          iterator_back=re.finditer(input_element[w][1],self.sequence)  
          p=2
          q=2
          sheet1 = wb1[input_name[w]]
          sheet2 = wb2[input_name[w]]
          wb1.save(cordence_result_xlsx) 
          wb2.save(discordence_result_xlsx) 
          t=sequence.find(input_element[w][1])

          spacer_num_width_w=input_spacer_number[w]
          snw=spacer_num_width_w
          len_input_element_w=len(input_element[w][0])
          lee=len_input_element_w

          if t==0:
             print(input_element[w][1]+" is not found")
             wb1.save(cordence_result_xlsx) 
             wb2.save(discordence_result_xlsx) 
             w=w+1 
          for n in iterator_back :
             e=n.group()
             f=n.start()
             g=n.end()
             print(e,f,g)
             if sequence[f-snw-lee:f-snw]==input_element[w][0]: 
                print("cordence sequence exists")          
                sheet1.cell(row=p, column=6, value=sequence[f-snw-lee:f-snw])
                sheet1.cell(row=p, column=7,value=sequence[f-snw:f])
                sheet1.cell(row=p, column=8, value=sequence[f:g])
                sheet1.cell(row=p, column=9, value=str(f-snw-lee+self.N_count)+":"+str(g+self.N_count))
                sheet1.cell(row=p, column=10, value=sequence.count(sequence[f-snw-lee:g]))
                p=p+1
                try: 
                   pass
                except IndexError as IE:
                   print(IE)
                   w=w+1            
                except PermissionError as PE:
                   print(PE)
                   continue
                else:
                   continue 
             elif sequence[f-snw-lee:f-snw]!=input_element[w][0]:
                #sheet2.cell(row=q, column=6, value=sequence[f-snw-lee:f-snw])
                #sheet2.cell(row=q, column=7,value=sequence[f-snw:f])
                #sheet2.cell(row=q, column=8, value=sequence[f:g])
                #sheet2.cell(row=q, column=9, value=str(f-snw-lee)+":"+str(g))
                #sheet2.cell(row=q, column=10, value=sequence.count(sequence[f-snw-lee:g]))
                q=q+1
                try: 
                   pass
                except IndexError as IE:
                   print(IE)
                   w=w+1            
                except PermissionError as PE:
                   print(PE)
                   continue
                else:
                   continue
             else:
                try: 
                   pass
                except IndexError as IE:
                   print(IE)
                   continue          
                except PermissionError as PE:
                   print(PE)
                   continue
                else:
                   continue
          else:
              wb1.save(cordence_result_xlsx) 
              wb2.save(discordence_result_xlsx)    
              w=w+1
         








  
 
